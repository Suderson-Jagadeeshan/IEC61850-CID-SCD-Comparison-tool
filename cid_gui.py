# ============================================================
#  IEC 61850 CID / SCD Comparison Tool
#  Version 2.0  –  Suderson Jagadeeshan
#
#  © 2024 Suderson Jagadeeshan. All rights reserved.
#  This software is proprietary and confidential.
#  Unauthorised copying, distribution or modification
#  is strictly prohibited.
#
#  Features:
#   - Auto technical key from IED name (no manual entry)
#   - CID and SCD support with IED selector
#   - Summary count at top of results
#   - Side-by-side diff view (Old | New)
#   - Progress bar during comparison
#   - Excel (.xlsx) export
#   - PDF export
#   - Colour-coded results
# ============================================================

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import datetime
import xml.etree.ElementTree as ET
import os
import threading

# ── Optional: openpyxl (Excel export) ───────────────────────
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ── Optional: reportlab (PDF export) ────────────────────────
try:
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.utils import simpleSplit
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    font_path = "DejaVuSans.ttf"
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
        PDF_FONT = 'DejaVuSans'
    else:
        PDF_FONT = 'Helvetica'
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    PDF_FONT = 'Helvetica'

# ── Optional: cid_compare_core ───────────────────────────────
try:
    from cid_compare_core import compare_cid_files
    CORE_AVAILABLE = True
except ImportError:
    CORE_AVAILABLE = False

SCL_NS = {'iec': 'http://www.iec.ch/61850/2003/SCL'}

# ════════════════════════════════════════════════════════════
#  LICENSE DIALOG
# ════════════════════════════════════════════════════════════

LICENSE_TEXT = """IEC 61850 CID / SCD Comparison Tool
Version 2.0

Copyright © 2024 Suderson Jagadeeshan
All rights reserved.

PROPRIETARY SOFTWARE NOTICE
─────────────────────────────────────────────────────
This software and its source code are the exclusive
property of Suderson Jagadeeshan.

You MAY:
  • Use this tool for internal engineering work
  • Generate reports for your projects

You MAY NOT:
  • Copy, redistribute or share this software
  • Modify or create derivative works
  • Use this software for commercial resale
  • Remove or alter this copyright notice

DISCLAIMER
─────────────────────────────────────────────────────
This tool is provided for engineering reference only.
Always verify comparison results against the original
IEC 61850 configuration files before making changes
to live substation systems.

The author accepts no liability for errors arising
from use of this tool in critical infrastructure.
─────────────────────────────────────────────────────
Contact: suderson123@gmail.com
"""

def show_license():
    win = tk.Toplevel(root)
    win.title("License & Copyright")
    win.resizable(False, False)
    win.grab_set()

    tk.Label(win, text="License & Copyright",
             font=("Arial", 12, "bold")).pack(pady=(12, 4))

    txt = tk.Text(win, width=62, height=30, wrap=tk.WORD,
                  font=("Courier", 9), bg="#F7F7F7", relief=tk.FLAT)
    txt.insert(tk.END, LICENSE_TEXT)
    txt.config(state=tk.DISABLED)
    txt.pack(padx=16, pady=4)

    tk.Button(win, text="Close", command=win.destroy,
              width=12).pack(pady=(4, 12))


# ════════════════════════════════════════════════════════════
#  HELPERS
# ════════════════════════════════════════════════════════════

def get_ied_names(filepath):
    try:
        tree = ET.parse(filepath)
        root = tree.getroot()
        ieds = root.findall('.//iec:IED', SCL_NS)
        if not ieds:
            ieds = root.findall('.//IED')
        return [ied.get('name', '(unnamed)') for ied in ieds]
    except Exception as e:
        print(f"IED list error: {e}")
        return []


def derive_prefix(ied_name):
    return ied_name.split('_')[0] if '_' in ied_name else ied_name


def on_ied_selected(event=None):
    ied_name = old_ied_combo.get()
    prefix   = derive_prefix(ied_name)
    tech_key_var.set(f"Technical Key: {prefix}   (IED: {ied_name})")


def browse_file(entry, ied_combo, is_old=False):
    path = filedialog.askopenfilename(
        filetypes=[("CID/SCD files", "*.cid *.scd"),
                   ("XML files",     "*.xml"),
                   ("All files",     "*.*")]
    )
    if not path:
        return
    entry.config(state=tk.NORMAL)
    entry.delete(0, tk.END)
    entry.insert(0, path)
    entry.config(state="readonly")

    ieds = get_ied_names(path)
    ied_combo['values'] = ieds
    ied_combo.set(ieds[0] if ieds else '(no IED found)')

    if is_old:
        on_ied_selected()
    validate_inputs()


def validate_inputs(*args):
    if old_entry.get().strip() and new_entry.get().strip():
        run_btn.config(state=tk.NORMAL)
    else:
        run_btn.config(state=tk.DISABLED)


def toggle_metadata(*args):
    mode = mode_var.get()
    # Enable metadata fields for both PDF and Excel, disable for View
    state = tk.NORMAL if mode in ('pdf', 'excel') else tk.DISABLED
    for w in metadata_widgets:
        w.config(state=state)


# ════════════════════════════════════════════════════════════
#  BUILT-IN COMPARISON
# ════════════════════════════════════════════════════════════

def parse_ied(filepath, ied_name=None):
    tree = ET.parse(filepath)
    root = tree.getroot()
    if ied_name and ied_name not in ('(no IED found)', '(unnamed)', ''):
        el = root.find(f'.//iec:IED[@name="{ied_name}"]', SCL_NS)
        if el is None:
            el = root.find(f'.//IED[@name="{ied_name}"]')
        return (el if el is not None else root), root
    return root, root


def builtin_compare(old_path, new_path, old_ied_name=None, new_ied_name=None):
    """
    Returns list of dicts:
      { 'category': str, 'element': str, 'old': str, 'new': str }
    """
    rows = []

    def row(cat, element, old_val, new_val):
        rows.append({'category': cat, 'element': element,
                     'old': str(old_val), 'new': str(new_val)})

    try:
        old_ied, old_full = parse_ied(old_path, old_ied_name)
        new_ied, new_full = parse_ied(new_path, new_ied_name)
    except Exception as e:
        return [{'category': 'ERROR', 'element': 'Parse error',
                 'old': str(e), 'new': ''}]

    def find_ied_el(root, name):
        el = root.find(f'.//iec:IED[@name="{name}"]', SCL_NS)
        return el if el is not None else root.find(f'.//IED[@name="{name}"]')

    # ── 1. IED attributes ─────────────────────────────────────
    old_el = find_ied_el(old_full, old_ied_name) if old_ied_name else None
    new_el = find_ied_el(new_full, new_ied_name) if new_ied_name else None
    if old_el is not None and new_el is not None:
        for attr in ['configVersion', 'desc', 'type', 'manufacturer']:
            ov, nv = old_el.get(attr,''), new_el.get(attr,'')
            if ov != nv:
                row('Config Revision', f"IED.{attr}", ov, nv)

    # ── 2. IP address ─────────────────────────────────────────
    def get_ip(root, ied_name):
        for tag in ['.//iec:ConnectedAP', './/ConnectedAP']:
            ns = SCL_NS if 'iec:' in tag else {}
            for cap in root.findall(tag, ns):
                if cap.get('iedName') == ied_name:
                    for p in list(cap.iter('{http://www.iec.ch/61850/2003/SCL}P')) + list(cap.iter('P')):
                        if p.get('type') == 'IP':
                            return p.text
        return 'N/A'

    old_ip = get_ip(old_full, old_ied_name)
    new_ip = get_ip(new_full, new_ied_name)
    if old_ip != new_ip:
        row('IP / Network', 'IP Address', old_ip, new_ip)

    # ── 3. DataSets ───────────────────────────────────────────
    def get_ds_names(ied_el):
        names = set()
        for tag in ['{http://www.iec.ch/61850/2003/SCL}DataSet', 'DataSet']:
            for ds in ied_el.iter(tag):
                names.add(ds.get('name',''))
        return names

    old_ds = get_ds_names(old_ied)
    new_ds = get_ds_names(new_ied)
    for n in old_ds - new_ds:
        row('DataSet', f"DataSet '{n}'", 'EXISTS', 'REMOVED')
    for n in new_ds - old_ds:
        row('DataSet', f"DataSet '{n}'", 'NOT PRESENT', 'ADDED')

    # ── 4. FCDAs ──────────────────────────────────────────────
    def get_fcdas(ied_el):
        result = {}
        for ns_tag in ['{http://www.iec.ch/61850/2003/SCL}DataSet', 'DataSet']:
            for ds in ied_el.iter(ns_tag):
                ds_name = ds.get('name','')
                for fc_tag in ['{http://www.iec.ch/61850/2003/SCL}FCDA', 'FCDA']:
                    for fc in ds.findall(fc_tag):
                        key = (ds_name, fc.get('ldInst',''), fc.get('lnClass',''),
                               fc.get('lnInst',''), fc.get('doName',''),
                               fc.get('daName',''), fc.get('fc',''))
                        result[key] = True
        return result

    old_fc = get_fcdas(old_ied)
    new_fc = get_fcdas(new_ied)
    for k in old_fc:
        if k not in new_fc:
            row('Signal / FCDA',
                f"DS '{k[0]}' → {k[2]}/{k[4]} fc={k[6]}",
                'EXISTS', 'REMOVED')
    for k in new_fc:
        if k not in old_fc:
            row('Signal / FCDA',
                f"DS '{k[0]}' → {k[2]}/{k[4]} fc={k[6]}",
                'NOT PRESENT', 'ADDED')

    # ── 5. ReportControl ──────────────────────────────────────
    def get_rcbs(ied_el):
        rcbs = {}
        for tag in ['{http://www.iec.ch/61850/2003/SCL}ReportControl', 'ReportControl']:
            for rc in ied_el.iter(tag):
                name = rc.get('name','')
                rcbs[name] = {a: rc.get(a,'') for a in
                              ['confRev','buffered','intgPd','datSet','rptID']}
        return rcbs

    old_rcbs = get_rcbs(old_ied)
    new_rcbs = get_rcbs(new_ied)
    for name, old_attrs in old_rcbs.items():
        if name not in new_rcbs:
            row('ReportControl', f"RCB '{name}'", 'EXISTS', 'REMOVED')
        else:
            for attr, ov in old_attrs.items():
                nv = new_rcbs[name].get(attr,'')
                if ov != nv:
                    row('ReportControl', f"RCB '{name}' → {attr}", ov, nv)
    for name in new_rcbs:
        if name not in old_rcbs:
            row('ReportControl', f"RCB '{name}'", 'NOT PRESENT', 'ADDED')

    # ── 6. Logical Nodes ──────────────────────────────────────
    def get_lns(ied_el):
        lns = set()
        for tag in ['{http://www.iec.ch/61850/2003/SCL}LN', 'LN']:
            for ln in ied_el.iter(tag):
                lns.add((ln.get('lnClass',''), ln.get('inst',''),
                         ln.get('ldInst',''), ln.get('desc','')))
        return lns

    old_lns = get_lns(old_ied)
    new_lns = get_lns(new_ied)
    for ln in old_lns - new_lns:
        row('Logical Node', f"{ln[0]} inst={ln[1]} LD={ln[2]}", 'EXISTS', 'REMOVED')
    for ln in new_lns - old_lns:
        row('Logical Node', f"{ln[0]} inst={ln[1]} LD={ln[2]}", 'NOT PRESENT', 'ADDED')

    return rows


# ════════════════════════════════════════════════════════════
#  SUMMARY & DISPLAY
# ════════════════════════════════════════════════════════════

CATEGORY_COLOURS = {
    'Config Revision': '#6A0DAD',   # purple
    'IP / Network':    '#8B4500',   # brown
    'DataSet':         '#00008B',   # dark blue
    'Signal / FCDA':   '#006400',   # dark green  (added)
    'ReportControl':   '#B8860B',   # dark goldenrod
    'Logical Node':    '#8B0000',   # dark red
    'ERROR':           '#FF0000',
}

REMOVE_COLOUR = '#8B0000'
ADD_COLOUR    = '#006400'


def build_summary(rows):
    from collections import Counter
    counts = Counter(r['category'] for r in rows)
    total  = len(rows)
    lines  = [f"  TOTAL DIFFERENCES FOUND: {total}  ",
              "  " + "─" * 44]
    for cat, cnt in sorted(counts.items()):
        lines.append(f"  {cat:<28} {cnt:>3} change{'s' if cnt != 1 else ''}")
    lines.append("  " + "─" * 44)
    return "\n".join(lines)


def display_results(rows):
    result_box.config(state=tk.NORMAL)
    result_box.delete(1.0, tk.END)

    # ── Configure tags ────────────────────────────────────────
    result_box.tag_config('summary',  background='#EEF2FF', font=("Courier", 9, "bold"))
    result_box.tag_config('header',   background='#D0D0D0', font=("Courier", 9, "bold"))
    result_box.tag_config('removed',  foreground=REMOVE_COLOUR)
    result_box.tag_config('added',    foreground=ADD_COLOUR)
    result_box.tag_config('changed',  foreground='#00008B')
    result_box.tag_config('ok',       foreground='#006400', font=("Courier", 9, "bold"))

    if not rows:
        result_box.insert(tk.END, "  ✔  No differences found.\n", 'ok')
        result_box.config(state=tk.DISABLED)
        return

    # ── Summary block ─────────────────────────────────────────
    result_box.insert(tk.END, build_summary(rows) + "\n\n", 'summary')

    # ── Side-by-side header ───────────────────────────────────
    col_w = 38
    header = f"  {'CATEGORY':<22}  {'ELEMENT':<35}  {'OLD VALUE':<{col_w}}  {'NEW VALUE':<{col_w}}\n"
    sep    = "  " + "─" * (22 + 35 + col_w * 2 + 10) + "\n"
    result_box.insert(tk.END, header, 'header')
    result_box.insert(tk.END, sep,    'header')

    # ── Rows ──────────────────────────────────────────────────
    for r in rows:
        old_v = r['old']
        new_v = r['new']

        if old_v == 'NOT PRESENT':
            tag = 'added'
        elif new_v == 'REMOVED':
            tag = 'removed'
        else:
            tag = 'changed'

        line = (f"  {r['category']:<22}  {r['element']:<35}  "
                f"{old_v:<{col_w}}  {new_v:<{col_w}}\n")
        result_box.insert(tk.END, line, tag)

    result_box.config(state=tk.DISABLED)


# ════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ════════════════════════════════════════════════════════════

def generate_excel_report(rows, metadata):
    if not EXCEL_AVAILABLE:
        messagebox.showerror("Excel Unavailable",
            "openpyxl is not installed.\n\nRun:\n  pip install openpyxl")
        return

    filename = f"CID_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = openpyxl.Workbook()

    # ── Styles ────────────────────────────────────────────────
    hdr_fill   = PatternFill("solid", fgColor="1F3864")
    hdr_font   = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    add_fill   = PatternFill("solid", fgColor="C6EFCE")
    rem_fill   = PatternFill("solid", fgColor="FFC7CE")
    chg_fill   = PatternFill("solid", fgColor="FFEB9C")
    sum_fill   = PatternFill("solid", fgColor="DCE6F1")
    thin       = Side(style='thin', color="AAAAAA")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_al  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_al    = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # ══ Sheet 1: Summary ══════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Summary"

    ws1.merge_cells('A1:D1')
    ws1['A1'] = "CID / SCD Comparison Report"
    ws1['A1'].font      = Font(bold=True, size=14, name="Calibri")
    ws1['A1'].alignment = center_al

    row_i = 2
    for label, value in metadata.items():
        ws1.cell(row=row_i, column=1, value=label).font = Font(bold=True, name="Calibri")
        ws1.cell(row=row_i, column=2, value=value)
        row_i += 1

    row_i += 1
    ws1.cell(row=row_i, column=1, value="Category").fill      = sum_fill
    ws1.cell(row=row_i, column=1).font                        = Font(bold=True, name="Calibri")
    ws1.cell(row=row_i, column=2, value="Count").fill         = sum_fill
    ws1.cell(row=row_i, column=2).font                        = Font(bold=True, name="Calibri")
    row_i += 1

    from collections import Counter
    counts = Counter(r['category'] for r in rows)
    for cat, cnt in sorted(counts.items()):
        ws1.cell(row=row_i, column=1, value=cat)
        ws1.cell(row=row_i, column=2, value=cnt)
        row_i += 1

    ws1.cell(row=row_i, column=1, value="TOTAL")
    ws1.cell(row=row_i, column=1).font = Font(bold=True, name="Calibri")
    ws1.cell(row=row_i, column=2, value=len(rows))
    ws1.cell(row=row_i, column=2).font = Font(bold=True, name="Calibri")

    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 18

    # ══ Sheet 2: Differences ══════════════════════════════════
    ws2 = wb.create_sheet("Differences")
    headers = ["#", "Category", "Element / Signal", "Old Value", "New Value", "Change Type"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = center_al
        cell.border    = border

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 45
    ws2.column_dimensions['D'].width = 30
    ws2.column_dimensions['E'].width = 30
    ws2.column_dimensions['F'].width = 14
    ws2.row_dimensions[1].height     = 20

    for i, r in enumerate(rows, 1):
        old_v = r['old']
        new_v = r['new']

        if old_v == 'NOT PRESENT':
            fill       = add_fill
            change_type = 'Added'
        elif new_v == 'REMOVED':
            fill       = rem_fill
            change_type = 'Removed'
        else:
            fill       = chg_fill
            change_type = 'Modified'

        vals = [i, r['category'], r['element'], old_v, new_v, change_type]
        for col, val in enumerate(vals, 1):
            cell            = ws2.cell(row=i+1, column=col, value=val)
            cell.fill       = fill
            cell.border     = border
            cell.alignment  = left_al
            cell.font       = Font(name="Calibri", size=10)

    # Freeze header row
    ws2.freeze_panes = 'A2'

    # ══ Sheet 3: Per-category tabs ════════════════════════════
    from collections import defaultdict
    by_cat = defaultdict(list)
    for r in rows:
        by_cat[r['category']].append(r)

    for cat, cat_rows in sorted(by_cat.items()):
        # Excel sheet names cannot contain: / \ ? * [ ] :  and max 31 chars
        safe_name = cat.replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '').replace('[', '').replace(']', '').replace(':', '')[:31]
        ws = wb.create_sheet(safe_name)
        for col, h in enumerate(["#", "Element / Signal", "Old Value", "New Value", "Change Type"], 1):
            cell            = ws.cell(row=1, column=col, value=h)
            cell.fill       = hdr_fill
            cell.font       = hdr_font
            cell.alignment  = center_al
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 14
        ws.freeze_panes = 'A2'

        for i, r in enumerate(cat_rows, 1):
            old_v = r['old']
            new_v = r['new']
            fill  = add_fill if old_v=='NOT PRESENT' else rem_fill if new_v=='REMOVED' else chg_fill
            ct    = 'Added' if old_v=='NOT PRESENT' else 'Removed' if new_v=='REMOVED' else 'Modified'
            for col, val in enumerate([i, r['element'], old_v, new_v, ct], 1):
                cell           = ws.cell(row=i+1, column=col, value=val)
                cell.fill      = fill
                cell.border    = border
                cell.alignment = left_al
                cell.font      = Font(name="Calibri", size=10)

    wb.save(filename)
    messagebox.showinfo("Excel Generated", f"Report saved as:\n{filename}")


# ════════════════════════════════════════════════════════════
#  PDF EXPORT
# ════════════════════════════════════════════════════════════

def draw_wrapped_text(c, x, y, text, max_width, font_name, font_size, page_height):
    if PDF_AVAILABLE:
        lines = simpleSplit(text, font_name, font_size, max_width)
    else:
        lines = [text]
    for line in lines:
        if y < 60:
            c.showPage()
            y = page_height - 50
            c.setFont(font_name, font_size)
        c.drawString(x, y, line)
        y -= font_size + 4
    return y


def generate_pdf_report(rows, metadata):
    if not PDF_AVAILABLE:
        messagebox.showerror("PDF Unavailable",
            "reportlab is not installed.\n\nRun:\n  pip install reportlab")
        return

    filename = f"CID_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    c = canvas.Canvas(filename, pagesize=A4)
    width, height = A4
    y = height - 50

    try:
        if os.path.exists("your_logo.png"):
            c.drawImage('your_logo.png', 50, y-35, width=80, height=35, mask='auto')
    except Exception as e:
        print(f"Logo error: {e}")
    y -= 100

    c.setFont(PDF_FONT, 14)
    c.drawString(160, y, "CID / SCD Comparison Report")
    y -= 30

    c.setFont(PDF_FONT, 10)
    for label, value in metadata.items():
        y = draw_wrapped_text(c, 50, y, f"{label}: {value}",
                              width-100, PDF_FONT, 10, height)
    y -= 10

    # Summary
    from collections import Counter
    counts = Counter(r['category'] for r in rows)
    c.setFont(PDF_FONT, 11)
    c.drawString(50, y, f"Total differences: {len(rows)}")
    y -= 18
    c.setFont(PDF_FONT, 9)
    for cat, cnt in sorted(counts.items()):
        c.drawString(70, y, f"{cat}: {cnt}")
        y -= 14
    y -= 10

    # Per-category sections
    from collections import defaultdict
    by_cat = defaultdict(list)
    for r in rows:
        by_cat[r['category']].append(r)

    for cat, cat_rows in sorted(by_cat.items()):
        if y < 100:
            c.showPage(); y = height - 50
        c.setFont(PDF_FONT, 11)
        c.drawString(50, y, f"[ {cat} ]")
        y -= 18
        c.setFont(PDF_FONT, 8)
        for r in cat_rows:
            line = f"  {r['element']}  |  {r['old']}  →  {r['new']}"
            y = draw_wrapped_text(c, 60, y, line, width-120, PDF_FONT, 8, height)
        y -= 8

    c.save()
    messagebox.showinfo("PDF Generated", f"Report saved as:\n{filename}")


# ════════════════════════════════════════════════════════════
#  RUN COMPARISON  (runs in background thread + progress bar)
# ════════════════════════════════════════════════════════════

def run_comparison():
    old_file = old_entry.get().strip()
    new_file = new_entry.get().strip()
    old_ied  = old_ied_combo.get().strip()
    new_ied  = new_ied_combo.get().strip()
    mode     = mode_var.get()

    if not old_file or not new_file:
        messagebox.showerror("Missing Input", "Please select both CID/SCD files.")
        return

    tech_prefix = derive_prefix(old_ied)

    # Lock UI, start progress bar
    run_btn.config(state=tk.DISABLED)
    progress_bar['value'] = 0
    progress_bar.grid()
    progress_label.grid()
    progress_label.config(text="Parsing files…")
    root.update_idletasks()

    def worker():
        try:
            progress_bar.after(0, lambda: setprog(20, "Comparing DataSets…"))

            if CORE_AVAILABLE:
                try:
                    raw = compare_cid_files(old_file, new_file, tech_prefix,
                                            old_ied=old_ied or None,
                                            new_ied=new_ied or None)
                    # Wrap plain strings into row dicts if core returns strings
                    if raw and isinstance(raw[0], str):
                        rows = [{'category': 'General', 'element': l,
                                 'old': '', 'new': ''} for l in raw]
                    else:
                        rows = raw
                except TypeError:
                    raw  = compare_cid_files(old_file, new_file, tech_prefix)
                    rows = [{'category': 'General', 'element': l,
                             'old': '', 'new': ''} for l in raw]
            else:
                rows = builtin_compare(old_file, new_file,
                                       old_ied_name=old_ied or None,
                                       new_ied_name=new_ied or None)

            progress_bar.after(0, lambda: setprog(80, "Rendering results…"))

            metadata = {
                "Old IED":       old_ied,
                "New IED":       new_ied,
                "Technical Key": tech_prefix,
                "Name":          name_entry.get(),
                "Designation":   designation_entry.get(),
                "Project Name":  project_entry.get(),
                "Voltage Level": voltage_entry.get(),
                "Report Date":   date_entry.get(),
                "Generated":     datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
            }

            def finish():
                setprog(100, "Done.")
                if mode == 'view':
                    display_results(rows)
                elif mode == 'pdf':
                    generate_pdf_report(rows, metadata)
                elif mode == 'excel':
                    generate_excel_report(rows, metadata)
                progress_bar.after(800, lambda: (
                    progress_bar.grid_remove(),
                    progress_label.grid_remove()
                ))
                run_btn.config(state=tk.NORMAL)

            progress_bar.after(0, finish)

        except Exception as e:
            def show_err():
                messagebox.showerror("Error", f"Comparison failed:\n{e}")
                progress_bar.grid_remove()
                progress_label.grid_remove()
                run_btn.config(state=tk.NORMAL)
            progress_bar.after(0, show_err)

    threading.Thread(target=worker, daemon=True).start()


def setprog(val, text):
    progress_bar['value'] = val
    progress_label.config(text=text)
    root.update_idletasks()


# ════════════════════════════════════════════════════════════
#  GUI LAYOUT
# ════════════════════════════════════════════════════════════

root = tk.Tk()
root.title("IEC 61850 CID / SCD Comparison Tool  v2.0")
root.resizable(True, True)

tech_key_var = tk.StringVar(value="Technical Key:  (select Old file first)")

# ── Menu bar ──────────────────────────────────────────────────
menubar = tk.Menu(root)
help_menu = tk.Menu(menubar, tearoff=0)
help_menu.add_command(label="License & Copyright", command=show_license)
menubar.add_cascade(label="Help", menu=help_menu)
root.config(menu=menubar)

# ── Logo & credit ─────────────────────────────────────────────
try:
    from PIL import Image, ImageTk
    img      = Image.open("your_logo.png").resize((100, 50))
    logo_img = ImageTk.PhotoImage(img)
    tk.Label(root, image=logo_img).grid(row=0, column=0, columnspan=5, pady=(10,0))
    tk.Label(root, text="Created by Suderson Jagadeeshan",
             font=("Arial", 10, "italic"), fg="gray").grid(row=1, column=0, columnspan=5)
except Exception:
    tk.Label(root, text="IEC 61850 CID / SCD Comparison Tool  v2.0",
             font=("Arial", 12, "bold")).grid(row=0, column=0, columnspan=5, pady=(10,0))
    tk.Label(root, text="Created by Suderson Jagadeeshan",
             font=("Arial", 10, "italic"), fg="gray").grid(row=1, column=0, columnspan=5)

# ── Column headers ────────────────────────────────────────────
tk.Label(root, text="File Path",  font=("Arial", 9, "bold")).grid(row=2, column=1, pady=(8,0))
tk.Label(root, text="Select IED", font=("Arial", 9, "bold")).grid(row=2, column=2, pady=(8,0))

# ── Old file ──────────────────────────────────────────────────
tk.Label(root, text="Old CID/SCD File").grid(row=3, column=0, sticky="e", padx=6, pady=4)
old_entry = tk.Entry(root, width=50, state="readonly")
old_entry.grid(row=3, column=1, padx=4)
old_ied_combo = ttk.Combobox(root, width=25, state="readonly")
old_ied_combo.grid(row=3, column=2, padx=4)
old_ied_combo.set("(browse first)")
old_ied_combo.bind("<<ComboboxSelected>>", on_ied_selected)
tk.Button(root, text="Browse",
          command=lambda: browse_file(old_entry, old_ied_combo, is_old=True)
          ).grid(row=3, column=3, padx=4)

# ── New file ──────────────────────────────────────────────────
tk.Label(root, text="New CID/SCD File").grid(row=4, column=0, sticky="e", padx=6, pady=4)
new_entry = tk.Entry(root, width=50, state="readonly")
new_entry.grid(row=4, column=1, padx=4)
new_ied_combo = ttk.Combobox(root, width=25, state="readonly")
new_ied_combo.grid(row=4, column=2, padx=4)
new_ied_combo.set("(browse first)")
tk.Button(root, text="Browse",
          command=lambda: browse_file(new_entry, new_ied_combo, is_old=False)
          ).grid(row=4, column=3, padx=4)

# ── Technical key display ─────────────────────────────────────
tk.Label(root, textvariable=tech_key_var, fg="#4A4A8A",
         font=("Arial", 9, "italic")).grid(row=5, column=0, columnspan=5, pady=(2,6))

# ── Output mode ───────────────────────────────────────────────
mode_var = tk.StringVar(value="view")
mode_var.trace_add('write', toggle_metadata)
tk.Label(root, text="Output Mode").grid(row=6, column=0, sticky="e", padx=6)
tk.Radiobutton(root, text="View Results",  variable=mode_var, value="view" ).grid(row=6, column=1, sticky="w")
tk.Radiobutton(root, text="Generate PDF",  variable=mode_var, value="pdf"  ).grid(row=6, column=2, sticky="w")
tk.Radiobutton(root, text="Export Excel",  variable=mode_var, value="excel").grid(row=6, column=3, sticky="w")

# ── PDF metadata (shown only in PDF mode) ─────────────────────
def lbl(text, row):
    tk.Label(root, text=text).grid(row=row, column=0, sticky="e", padx=6, pady=2)

lbl("Name",          7);  name_entry        = tk.Entry(root, width=35, state=tk.DISABLED); name_entry.grid(row=7,  column=1, sticky="w", padx=4)
lbl("Designation",   8);  designation_entry = tk.Entry(root, width=35, state=tk.DISABLED); designation_entry.grid(row=8,  column=1, sticky="w", padx=4)
lbl("Project Name",  9);  project_entry     = tk.Entry(root, width=35, state=tk.DISABLED); project_entry.grid(row=9,  column=1, sticky="w", padx=4)
lbl("Voltage Level", 10); voltage_entry     = tk.Entry(root, width=35, state=tk.DISABLED); voltage_entry.grid(row=10, column=1, sticky="w", padx=4)
lbl("Report Date",   11); date_entry        = tk.Entry(root, width=35, state=tk.DISABLED); date_entry.grid(row=11, column=1, sticky="w", padx=4)

date_entry.config(state=tk.NORMAL)
date_entry.insert(0, datetime.datetime.now().strftime('%Y-%m-%d'))
date_entry.config(state=tk.DISABLED)

pdf_widgets   = [name_entry, designation_entry, project_entry, voltage_entry, date_entry]
excel_widgets = [name_entry, designation_entry, project_entry, voltage_entry, date_entry]
metadata_widgets = pdf_widgets   # kept for compat

# ── Run button ────────────────────────────────────────────────
run_btn = tk.Button(root, text="▶  Run Comparison", command=run_comparison,
                    state=tk.DISABLED, width=22,
                    bg="#4CAF50", fg="white", font=("Arial", 10, "bold"))
run_btn.grid(row=12, column=1, pady=10)

# ── Progress bar ──────────────────────────────────────────────
progress_bar = ttk.Progressbar(root, orient=tk.HORIZONTAL,
                                length=400, mode='determinate')
progress_bar.grid(row=13, column=0, columnspan=4, padx=20, pady=(0,2))
progress_bar.grid_remove()   # hidden until comparison starts

progress_label = tk.Label(root, text="", fg="#4A4A8A", font=("Arial", 8, "italic"))
progress_label.grid(row=14, column=0, columnspan=4)
progress_label.grid_remove()

# ── Results area ──────────────────────────────────────────────
tk.Label(root, text="Comparison Results  (colour: green=added, red=removed, blue=modified)",
         font=("Arial", 9, "bold"), fg="#333333").grid(
    row=15, column=0, columnspan=5, sticky="w", padx=12, pady=(4,0))

result_frame = tk.Frame(root)
result_frame.grid(row=16, column=0, columnspan=5, padx=10, pady=(0,6), sticky="nsew")
root.grid_rowconfigure(16, weight=1)
root.grid_columnconfigure(1, weight=1)

result_box = tk.Text(result_frame, width=120, height=24, wrap=tk.NONE,
                     state=tk.DISABLED, font=("Courier", 9))
vsb = tk.Scrollbar(result_frame, orient=tk.VERTICAL,   command=result_box.yview)
hsb = tk.Scrollbar(result_frame, orient=tk.HORIZONTAL, command=result_box.xview)
result_box.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
result_box.grid(row=0, column=0, sticky="nsew")
vsb.grid(row=0, column=1, sticky="ns")
hsb.grid(row=1, column=0, sticky="ew")
result_frame.grid_rowconfigure(0, weight=1)
result_frame.grid_columnconfigure(0, weight=1)

# ── Status bar ────────────────────────────────────────────────
status_parts = ["Ready  –  v2.0  © 2024 Suderson Jagadeeshan"]
if not CORE_AVAILABLE:
    status_parts.append("built-in comparison active")
if not PDF_AVAILABLE:
    status_parts.append("PDF disabled (pip install reportlab)")
if not EXCEL_AVAILABLE:
    status_parts.append("Excel disabled (pip install openpyxl)")

tk.Label(root, text="   |   ".join(status_parts), fg="gray",
         font=("Arial", 8), anchor="w").grid(
    row=17, column=0, columnspan=5, sticky="ew", padx=10, pady=(0,4))

root.mainloop()
